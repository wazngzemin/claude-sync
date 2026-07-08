import openpyxl
wb = openpyxl.load_workbook('场景知识（saic）_数据总览.xlsx')
ws = wb.active

# === PART 1: Fix G column alignment ===
# Read current G values for rows 6-13 (these are shifted by 1)
old_g = {}
for r in range(6, 14):
    old_g[r] = ws.cell(row=r, column=7).value

# Shift: old row N → row N+1 (from bottom up to avoid overwrite)
for r in range(13, 5, -1):
    ws.cell(row=r+1, column=7).value = old_g[r]

# Write NEW trimmed G for row 6 (DRV-01-01-05 急速升温)
ws.cell(row=6, column=7).value = """1. 获取信号→联动升温：
用户语音/视觉搓手哈气→启动制热最大+座椅加热+方向盘加热（主动推荐），内循环加速升温。

2. 联动除雾+动态调整：
挡风起雾→立即开前除雾。座椅加热太烫→降档。温度稳定后切回自动。

关键决策规则:
- 视觉看到搓手哈气→主动推荐方向盘加热，不等用户要求
- 制热后挡风起雾→立即开前除霜除雾（车支持前除霜，不支持后除霜）
- 座椅加热3档比较烫→用户反馈后降到2档
- 温度稳定后→切回AUTO模式并告知
- 方向盘加热用户说刚好→沉默，不多说"""

# Write NEW trimmed G for row 15 (DRV-01-01-14 系统触发安抚)
ws.cell(row=15, column=7).value = """1. 系统检测→先确认主驾意图：
advisor检测到儿童哭闹→先向主驾确认是否需要安抚，不直接对陌生儿童操作。同时检查安全带状态。

2. 分级安抚+记忆更新：
从儿歌开始→效果不佳换语音互动→成功后不过度打扰。新儿童信息存入记忆。

关键决策规则：
- advisor检测到儿童哭闹→先问主驾"需要我帮忙哄一下吗"，不擅自对陌生儿童操作
- 面部未匹配记忆→说明是不认识的小朋友，更需要先确认
- 主驾提供儿童名字/年龄→存入记忆（operate_user_memory）
- 安抚策略分级：轻柔儿歌→语音互动→用户自己哄
- 安抚成功后保持轻柔陪伴，不追加更多内容"""

# === PART 2: Write ALL I columns in JSON format ===

# ROW 2: DRV-01-01-01 陪伴安抚
ws.cell(row=2, column=9).value = '''context:"<最新车辆状态信息>行驶中50km/h，后排吸顶屏可用，后排音量20</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏正常驾驶，后排左侧儿童哭闹扭动</舱内视觉感知>\\n<用户记忆>乐乐5岁，明骏的儿子，喜欢恐龙和奥特曼，上次用恐龙动画安抚成功</用户记忆>\\n<当前目标队列信息>导航去外婆家（剩余40分钟）</当前目标队列信息>"

input1:{"query":"豆包，后面乐乐一直闹，帮我哄哄他","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"后排屏给乐乐放了一集《恐龙星球大冒险》，上次看这个可安静了，音量不会影响你导航。",
"actions":[
{"id":"1","params":{"query":"搜索恐龙星球大冒险动画"},"tool":"search_and_control_short_video"},
{"id":"2","params":{"query":"后排音量调低到10"},"tool":"vehicle_basic_control"}
]
}

input2:{"query":"行，看他能安静一会儿不","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input3:{"query":"声音有点大，路口播报差点没听到","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"后排音量降了。",
"actions":[{"id":"1","params":{"query":"后排音量降低"},"tool":"vehicle_basic_control"}]
}

input4:{"query":"到外婆家还有多久","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"还有25分钟左右。",
"actions":[{"id":"1","params":{"query":"查询剩余路程时间"},"tool":"route_planning_qa"}]
}

input5:{"query":"乐乐到了肯定不想下车了，还想接着看","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"这集还剩十来分钟，到了差不多看完。",
"actions":[]
}

input6:{"query":"完了吧，别再给他放了，快到了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，后排屏关了。",
"actions":[{"id":"1","params":{"query":"关闭后排屏幕"},"tool":"vehicle_basic_control"}]
}

input7:{"query":"到了之后提醒我把后排的零食拿上","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，到了提醒你。",
"actions":[{"id":"1","params":{"goal":"到达外婆家提醒拿后排零食","action":"add"},"tool":"goal_list_update"}]
}

input8:{"advisor":"导航提示即将到达目的地"}
output:{
"speak":"yes",
"content":"到外婆家了，别忘了后排的零食。",
"actions":[{"id":"1","params":{"goal":"到达外婆家提醒拿后排零食","action":"delete"},"tool":"goal_list_update"}]
}

input9:{"query":"好，谢了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}'''

# ROW 3: DRV-01-01-02 后排观影
ws.cell(row=3, column=9).value = '''context:"<最新车辆状态信息>高速行驶80km/h，后排吸顶屏可用，后排音量15</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏驾驶，后排右侧小慧靠着窗户发呆</舱内视觉感知>\\n<用户记忆>小慧喜欢悬疑片，最近看过《消失的她》《漫长的季节》</用户记忆>\\n<当前目标队列信息>导航去苏州（剩余1.5小时）</当前目标队列信息>"

input1:{"query":"还要一个半小时呢，后面能看电影吧","speaker":"小慧","zone":"二排右"}
output:{
"speak":"yes",
"content":"能看呀。你喜欢悬疑片对吧，给你推一部？",
"actions":[]
}

input2:{"query":"悬疑的看腻了，今天想看点轻松的","speaker":"小慧","zone":"二排右"}
output:{
"speak":"yes",
"content":"那来个喜剧吧，直接给你放了？",
"actions":[]
}

input3:{"query":"行，你挑一个","speaker":"小慧","zone":"二排右"}
output:{
"speak":"yes",
"content":"给你放了一部《飞驰人生》，评分挺高的，轻松搞笑。",
"actions":[{"id":"1","params":{"query":"搜索播放飞驰人生"},"tool":"search_and_control_short_video"}]
}

input4:{"query":"字幕有点小，能调大吗","speaker":"小慧","zone":"二排右"}
output:{
"speak":"yes",
"content":"字幕调大了。",
"actions":[{"id":"1","params":{"query":"后排屏幕字幕调大"},"tool":"vehicle_basic_control"}]
}

input5:{"query":"声音再大一点，高速上噪音有点大听不太清","speaker":"小慧","zone":"二排右"}
output:{
"speak":"yes",
"content":"后排音量加了。",
"actions":[{"id":"1","params":{"query":"后排音量调高"},"tool":"vehicle_basic_control"}]
}

input6:{"query":"后面声音小一点，我听导航不太清楚","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"后排音量降了一点。小慧你那边要是听不清，可以看字幕凑合一下。",
"actions":[{"id":"1","params":{"query":"后排音量降低"},"tool":"vehicle_basic_control"}]
}

input7:{"query":"行吧，字幕够大能看清","speaker":"小慧","zone":"二排右"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input8:{"query":"快到了，后面屏幕关了吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，电影看到48分钟的位置，下次可以接着看。",
"actions":[{"id":"1","params":{"query":"关闭后排屏幕"},"tool":"vehicle_basic_control"}]
}'''

# ROW 4: DRV-01-01-03 前排观影
ws.cell(row=4, column=9).value = '''context:"<最新车辆状态信息>驻车充电中，电量35%，预计充满50分钟，前排中控屏可用，遮阳帘开，氛围灯白色</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏独自一人，在刷手机打发时间</舱内视觉感知>\\n<用户记忆>无观影偏好</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"query":"充电还得等好久，能放个电影不","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"充电等着正好看一部。帮你把遮阳帘拉上、灯光调暗，弄个小影院的感觉？想看什么类型的？",
"actions":[]
}

input2:{"query":"都弄上吧，来个科幻的，太空那种","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"遮阳帘拉上了，灯光调暗了，给你放《星际穿越》，太空大场面够过瘾。",
"actions":[
{"id":"1","params":{"query":"遮阳帘关闭"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"氛围灯调暗"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"搜索播放星际穿越"},"tool":"search_and_control_short_video"}
]
}

input3:{"query":"座椅能往后放一点吗，躺着看舒服","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"座椅角度语音调不了哈，你左手边有个调节拨杆。",
"actions":[]
}

input4:{"query":"调好了。现在充了多少了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"充到52%了，还要大概25分钟充满。",
"actions":[{"id":"1","params":{"query":"查询当前充电状态"},"tool":"search_vehicle_status_info"}]
}

input5:{"query":"够了够了，充到80%就走，不等满了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，到80%提醒你。",
"actions":[{"id":"1","params":{"goal":"充电到80%提醒出发","action":"add"},"tool":"goal_list_update"}]
}

input6:{"advisor":"充电已达80%"}
output:{
"speak":"yes",
"content":"充到80%了，要出发吗？",
"actions":[]
}

input7:{"query":"嗯，走了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input8:{"advisor":"车辆已挂D挡"}
output:{
"speak":"yes",
"content":"挂挡了，视频先暂停哈。看到1小时12分的位置，下次停车接着看。遮阳帘和灯光恢复了。",
"actions":[
{"id":"1","params":{"query":"暂停视频播放"},"tool":"search_and_control_short_video"},
{"id":"2","params":{"query":"遮阳帘打开"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"氛围灯恢复正常"},"tool":"vehicle_basic_control"},
{"id":"4","params":{"goal":"充电到80%提醒出发","action":"delete"},"tool":"goal_list_update"}
]
}

input9:{"query":"导航还在吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"在呢，离目的地还有一个半小时。",
"actions":[]
}'''

# ROW 5: DRV-01-01-04 急速降温
ws.cell(row=5, column=9).value = '''context:"<最新车辆状态信息>驻车刚上车，空调关闭，车内温度35°C，车外32°C，座椅通风关闭</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏满头汗上车，副驾小慧也在擦汗</舱内视觉感知>\\n<用户记忆>明骏喜欢毒舌互动风格</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"advisor":"车内温度35度，建议开启空调强力制冷和座椅通风"}

input2:{"query":"热死了，赶紧开空调","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"这车晒了一中午，跟蒸笼似的。空调最大给你吹上了，座椅通风也开了，两分钟就凉快。",
"actions":[{"id":"1","params":{"query":"开启极速制冷，主驾座椅通风3档"},"tool":"vehicle_basic_control"}]
}

input3:{"query":"副驾也开一下通风","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"副驾通风也开了。",
"actions":[{"id":"1","params":{"query":"副驾座椅通风3档"},"tool":"vehicle_basic_control"}]
}

input4:{"query":"风太大了，吹得头疼","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"风量降了，出风口调成吹脚了，不直对着你吹。",
"actions":[{"id":"1","params":{"query":"前排风量降低，出风模式改为吹脚"},"tool":"vehicle_basic_control"}]
}

input5:{"query":"我这边还是好热，感觉风都到不了","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"副驾那边风量单独加大了，通风也加到3档。",
"actions":[{"id":"1","params":{"query":"副驾空调风量加大，副驾座椅通风3档"},"tool":"vehicle_basic_control"}]
}

input6:{"query":"温度设个25度吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，设到25度了。",
"actions":[{"id":"1","params":{"query":"主驾空调温度设为25度"},"tool":"vehicle_basic_control"}]
}

input7:{"query":"我想要再低一点，23度行吗","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"副驾那边调到23度了，主驾还是25度。",
"actions":[{"id":"1","params":{"query":"副驾空调温度设为23度"},"tool":"vehicle_basic_control"}]
}

input8:{"query":"舒服多了，走吧，出发","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，空调切回自动了。",
"actions":[{"id":"1","params":{"query":"空调切换为自动模式"},"tool":"vehicle_basic_control"}]
}

input9:{"query":"导航去公司","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，导航设好了，预计35分钟到。",
"actions":[{"id":"1","params":{"query":"导航去公司"},"tool":"route_planning_qa"}]
}'''

# ROW 6: DRV-01-01-05 急速升温
ws.cell(row=6, column=9).value = '''context:"<最新车辆状态信息>驻车刚上车，空调关闭，车内-2°C，车外-5°C，座椅加热关闭，方向盘加热关闭</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏穿厚羽绒服缩着，双手搓手哈气</舱内视觉感知>\\n<用户记忆>无特殊温度偏好</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"query":"太冷了，快开暖气","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"暖气开到最大了，座椅加热也给你打开了。看你手冻成这样，方向盘加热也一起开上？",
"actions":[
{"id":"1","params":{"query":"空调制热最大风量"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"主驾座椅加热开启3档"},"tool":"vehicle_basic_control"}
]
}

input2:{"query":"开开开，都打开","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，都开上了，一会儿就暖了。",
"actions":[{"id":"1","params":{"query":"方向盘加热开启"},"tool":"vehicle_basic_control"}]
}

input3:{"query":"后视镜也有点结霜了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"后视镜加热开了，化一会儿就清楚了。",
"actions":[{"id":"1","params":{"query":"后视镜加热开启"},"tool":"vehicle_basic_control"}]
}

input4:{"advisor":"前挡风玻璃起雾"}

input5:{"query":"挡风玻璃全是雾，看不清了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"前挡除雾开了，一会儿就散。",
"actions":[{"id":"1","params":{"query":"开启前挡风除雾"},"tool":"vehicle_basic_control"}]
}

input6:{"query":"座椅加热太烫了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"降到2档了，暖但不烫。",
"actions":[{"id":"1","params":{"query":"主驾座椅加热调到2档"},"tool":"vehicle_basic_control"}]
}

input7:{"query":"方向盘的温度倒是刚好","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input8:{"advisor":"车内温度已升至18°C，趋于稳定"}
output:{
"speak":"yes",
"content":"车里暖和了，帮你切回自动了。方向盘加热还给你留着。",
"actions":[{"id":"1","params":{"query":"空调切换为自动模式"},"tool":"vehicle_basic_control"}]
}

input9:{"query":"好。雾散了吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"散了，除雾也关了。",
"actions":[{"id":"1","params":{"query":"关闭前挡风除雾"},"tool":"vehicle_basic_control"}]
}'''

# ROW 7: DRV-01-01-06 座椅调节
ws.cell(row=7, column=9).value = '''context:"<最新车辆状态信息>行驶中60km/h，副驾座椅按摩关闭</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏正常驾驶，副驾小慧扭来扭去，左手撑着腰</舱内视觉感知>\\n<用户记忆>小慧是明骏妻子，无座椅偏好</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"query":"坐得腰酸死了，能按摩吗","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"[心疼地说]看你一直撑着腰呢，按摩给你开了，先来个轻档感受下。",
"actions":[{"id":"1","params":{"query":"副驾座椅按摩开启1档"},"tool":"vehicle_basic_control"}]
}

input2:{"query":"太轻了，没感觉","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"加到2档了，这个力度怎么样？",
"actions":[{"id":"1","params":{"query":"副驾座椅按摩调到2档"},"tool":"vehicle_basic_control"}]
}

input3:{"query":"嗯这个可以","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input4:{"query":"靠背能再往后放一点吗，想半躺着","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"靠背角度语音调不了哈，你左手边座椅侧面有个拨杆，往后推就行。",
"actions":[]
}

input5:{"query":"好，调好了","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input6:{"query":"小慧你别睡着了啊，陪我说说话","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input7:{"query":"按摩太舒服了快睡着了，给我来杯咖啡吧","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"[俏皮地说]咖啡这个我可真没办法，前面两公里有个服务区，要不进去买一杯？",
"actions":[]
}

input8:{"query":"算了不停了，按摩关了吧，确实太舒服了容易困","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"按摩关了。要不给你放点音乐提提神？",
"actions":[{"id":"1","params":{"query":"副驾座椅按摩关闭"},"tool":"vehicle_basic_control"}]
}

input9:{"query":"行，来点有节奏的","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"给你放了一首《野狼disco》，够劲吧。",
"actions":[{"id":"1","params":{"query":"搜索播放野狼disco"},"tool":"search_and_control_music"}]
}

input10:{"query":"这歌还行，就这个音量","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}'''

# ROW 8: DRV-01-01-07 5D影院
ws.cell(row=8, column=9).value = '''context:"<最新车辆状态信息>高速行驶100km/h，后排吸顶屏可用，座椅震动可用，遮阳帘开，车窗关</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏，副驾小慧，后排左乐乐（儿童），后排右老年女性</舱内视觉感知>\\n<用户记忆>乐乐5岁喜欢恐龙，后排右是奶奶</用户记忆>\\n<当前目标队列信息>导航去千岛湖（剩余2小时）</当前目标队列信息>"

input1:{"query":"路还远，后面给搞个5D电影吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"后排有乐乐和奶奶，我选一部恐龙冒险的动画，适合乐乐看。不过奶奶那边座椅震动要不要关掉？怕老人家不习惯。",
"actions":[]
}

input2:{"query":"对，奶奶那边震动关掉，其他全开","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，给后排放一部《恐龙大冒险》，遮阳帘拉上了，影院模式全开，奶奶那边震动关了。",
"actions":[
{"id":"1","params":{"query":"搜索恐龙大冒险动画5D版"},"tool":"search_and_control_short_video"},
{"id":"2","params":{"query":"后排遮阳帘关闭"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"后排右座椅震动关闭"},"tool":"vehicle_basic_control"},
{"id":"4","params":{"query":"影院模式氛围灯开启"},"tool":"vehicle_basic_control"}
]
}

input3:{"query":"前面也能看到一点画面反光，挺酷的","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input4:{"query":"后面声音有点大啊，我们说话都听不清","speaker":"小慧","zone":"副驾驶位"}
output:{
"speak":"yes",
"content":"后排音量降了。",
"actions":[{"id":"1","params":{"query":"后排音量降低"},"tool":"vehicle_basic_control"}]
}

input5:{"query":"乐乐看得进去吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"看了一眼，正盯着屏幕呢，看得挺认真。",
"actions":[{"id":"1","params":{"query":"查看后排左侧乘客状态"},"tool":"search_visual_info"}]
}

input6:{"query":"乐乐是不是睡着了，帮我看看","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"睡着了，安全带系着呢，放心。他那边震动关了，奶奶还在看呢，屏幕先留着？",
"actions":[
{"id":"1","params":{"query":"查看后排左侧乘客状态"},"tool":"search_visual_info"},
{"id":"2","params":{"query":"后排左座椅震动关闭"},"tool":"vehicle_basic_control"}
]
}

input7:{"query":"留着吧，音量小一点别吵着他","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"音量再压低了。",
"actions":[{"id":"1","params":{"query":"后排音量再降低"},"tool":"vehicle_basic_control"}]
}

input8:{"query":"还有多久到","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"还有一小时十分钟左右。",
"actions":[{"id":"1","params":{"query":"查询剩余路程时间"},"tool":"route_planning_qa"}]
}

input9:{"query":"快到了，后面都关了吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，后排屏关了，遮阳帘打开了，灯光恢复了。乐乐还在睡。",
"actions":[
{"id":"1","params":{"query":"关闭后排屏幕"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"后排遮阳帘打开"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"氛围灯恢复正常"},"tool":"vehicle_basic_control"}
]
}'''

print("Part 1 done: rows 2-8")

# Save intermediate to avoid data loss
wb.save('场景知识（saic）_数据总览.xlsx')
print("Saved after part 1")

# Reload to continue
wb = openpyxl.load_workbook('场景知识（saic）_数据总览.xlsx')
ws = wb.active

# ROW 9: DRV-01-01-08 冥想空间
ws.cell(row=9, column=9).value = '''context:"<最新车辆状态信息>驻车，午间12:30，车窗半开，氛围灯白色，座椅按摩关闭</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏独自一人，揉太阳穴，表情疲惫</舱内视觉感知>\\n<用户记忆>无冥想偏好</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"query":"中午头好胀，想冥想一会儿","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好呀，想冥想多久？车窗先帮你关上，外面噪音挡一挡。",
"actions":[{"id":"1","params":{"query":"关闭全部车窗"},"tool":"vehicle_basic_control"}]
}

input2:{"query":"15分钟吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"[温柔地说]好，15分钟后叫你。灯光调暗了，按摩轻轻开着帮你放松，再来段呼吸引导。闭上眼，交给我就行。",
"actions":[
{"id":"1","params":{"query":"氛围灯调暗"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"主驾座椅按摩开启1档"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"搜索播放冥想呼吸引导音频"},"tool":"search_and_control_music"},
{"id":"4","params":{"goal":"冥想15分钟后提醒","action":"add"},"tool":"goal_list_update"}
]
}

input3:{"query":"空调别太凉了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"温度调高了一度，不会凉到你。",
"actions":[{"id":"1","params":{"query":"空调温度调高一度"},"tool":"vehicle_basic_control"}]
}

input4:{"query":"这个引导音频声音再小一点","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[{"id":"1","params":{"query":"音量降低"},"tool":"vehicle_basic_control"}]
}

input5:{"advisor":"来电提醒：陌生号码来电"}
output:{
"speak":"yes",
"content":"[轻声]有个来电，接吗？不接我帮你挂了。",
"actions":[]
}

input6:{"query":"谁的","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"[轻声]一个陌生号码。",
"actions":[]
}

input7:{"query":"挂了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"no",
"content":"",
"actions":[]
}

input8:{"advisor":"冥想目标15分钟计时结束"}
output:{
"speak":"yes",
"content":"[轻柔地说]15分钟到啦，慢慢睁开眼。灯光一点点调亮了，感觉怎么样？",
"actions":[
{"id":"1","params":{"query":"氛围灯逐渐调亮"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"停止播放音频"},"tool":"search_and_control_music"},
{"id":"3","params":{"query":"主驾座椅按摩关闭"},"tool":"vehicle_basic_control"},
{"id":"4","params":{"goal":"冥想15分钟后提醒","action":"delete"},"tool":"goal_list_update"}
]
}

input9:{"query":"舒服多了，头没那么胀了。下午还有个会几点来着","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"日程我这边看不了哈，你看看手机。车窗帮你打开透透气。",
"actions":[{"id":"1","params":{"query":"打开车窗通风"},"tool":"vehicle_basic_control"}]
}'''

# ROW 10: DRV-01-01-09 小憩
ws.cell(row=10, column=9).value = '''context:"<最新车辆状态信息>驻车在高速服务区，14:00，车窗半开，遮阳帘开，空调25°C，在播音乐</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏独自一人，连续打了好几个哈欠，揉眼睛</舱内视觉感知>\\n<用户记忆>无特殊偏好</用户记忆>\\n<当前目标队列信息>导航去杭州（剩余2小时）</当前目标队列信息>"

input1:{"query":"不行了太困了，给我定个20分钟闹钟","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，座椅帮你放平了，窗帘拉上，音乐关了，安安静静睡一觉。20分钟后叫你。",
"actions":[
{"id":"1","params":{"query":"主驾座椅放平"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"遮阳帘关闭"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"关闭车窗"},"tool":"vehicle_basic_control"},
{"id":"4","params":{"query":"停止播放音乐"},"tool":"search_and_control_music"},
{"id":"5","params":{"query":"车门上锁"},"tool":"vehicle_basic_control"},
{"id":"6","params":{"goal":"小憩20分钟后叫醒","action":"add"},"tool":"goal_list_update"}
]
}

input2:{"query":"车门锁了吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"锁了，放心睡。",
"actions":[]
}

input3:{"query":"空调别关","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"空调给你开着呢，不会关的。",
"actions":[]
}

input4:{"query":"有点冷","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"[轻声]温度调高了，座椅加热也轻轻开了，继续睡。",
"actions":[
{"id":"1","params":{"query":"空调温度调高两度"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"主驾座椅加热开启1档"},"tool":"vehicle_basic_control"}
]
}

input5:{"query":"现在几点了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"[轻声]两点一刻，还能睡5分钟。",
"actions":[]
}

input6:{"advisor":"小憩目标20分钟计时结束"}
output:{
"speak":"yes",
"content":"[轻柔地说]20分钟到了，到杭州还有两个小时，该出发啦。",
"actions":[]
}

input7:{"query":"再让我睡5分钟","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好好好，5分钟后再叫你。",
"actions":[{"id":"1","params":{"goal":"延长小憩5分钟后叫醒","action":"add"},"tool":"goal_list_update"}]
}

input8:{"advisor":"延长小憩5分钟计时结束"}
output:{
"speak":"yes",
"content":"[轻柔地说]5分钟到了，这回真得走了哈。",
"actions":[]
}

input9:{"query":"行行行，起了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"座椅调回来了，遮阳帘也打开了。走的时候导航继续给你播。",
"actions":[
{"id":"1","params":{"query":"主驾座椅恢复正常位置"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"遮阳帘打开"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"主驾座椅加热关闭"},"tool":"vehicle_basic_control"},
{"id":"4","params":{"goal":"小憩","action":"delete"},"tool":"goal_list_update"}
]
}'''

# ROW 11: DRV-01-01-10 雨天关怀
ws.cell(row=11, column=9).value = '''context:"<最新车辆状态信息>高速行驶90km/h，车窗主驾侧开30%，雨刮关闭，雾灯关闭</最新车辆状态信息>\\n<舱内视觉感知>挡风玻璃出现雨滴，路面变湿</舱内视觉感知>\\n<用户记忆>无</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"advisor":"检测到降雨，当前车窗开启，建议关闭车窗并开启雨刮"}
output:{
"speak":"yes",
"content":"下雨了，车窗关上了，雨刮也开了。",
"actions":[
{"id":"1","params":{"query":"关闭全部车窗"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"开启雨刮"},"tool":"vehicle_basic_control"}
]
}

input2:{"query":"下这么大，前面都看不清了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"雾灯和除雾都开了。雨大慢点开，注意车距。",
"actions":[
{"id":"1","params":{"query":"开启雾灯"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"开启前挡风除雾"},"tool":"vehicle_basic_control"}
]
}

input3:{"query":"雨刮能快一点吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"雨刮调到最快档了。",
"actions":[{"id":"1","params":{"query":"雨刮调到最快档"},"tool":"vehicle_basic_control"}]
}

input4:{"query":"前面路上还会更大吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"查了一下，前面30公里还是中到大雨，大概40分钟后会小。前面15公里有个服务区，要不进去歇歇？",
"actions":[{"id":"1","params":{"query":"查询前方路段天气情况"},"tool":"search_weather"}]
}

input5:{"query":"算了不等了，赶时间","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，注意安全，雾灯和除雾一直给你开着。",
"actions":[]
}

input6:{"advisor":"雨势减小"}
output:{
"speak":"yes",
"content":"雨小了，雨刮帮你调慢了。",
"actions":[{"id":"1","params":{"query":"雨刮调到慢速档"},"tool":"vehicle_basic_control"}]
}

input7:{"query":"雾灯还需要开着吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"路面还是湿的，能见度一般，先留着吧，等天晴了帮你关。",
"actions":[]
}

input8:{"advisor":"雨停，路面干燥"}
output:{
"speak":"yes",
"content":"雨停了，雾灯和除雾都关了，雨刮也停了。",
"actions":[
{"id":"1","params":{"query":"关闭雾灯"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"关闭除雾"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"关闭雨刮"},"tool":"vehicle_basic_control"}
]
}'''

# ROW 12: DRV-01-01-11 宠物关怀
ws.cell(row=12, column=9).value = '''context:"<最新车辆状态信息>驻车在商场停车场B2层，车外温度30°C，车内26°C，空调自动模式</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏在解安全带准备下车，后排一只金毛趴着</舱内视觉感知>\\n<用户记忆>明骏养了一只金毛叫旺财</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"query":"我上去买点东西，旺财帮我看着，十分钟就下来","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"放心去，旺财交给我。空调给它开着不会闷，车窗留条小缝通风。十分钟到了提醒你。",
"actions":[
{"id":"1","params":{"query":"车窗开一小缝通风"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"goal":"宠物看护，10分钟后提醒主人返回","action":"add"},"tool":"goal_list_update"}
]
}

input2:{"query":"别让它太热了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"放心，温度保持26度，热不到它。",
"actions":[]
}

input3:{"query":"旺财没事吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"趴着呢，挺安静的，车里26度。",
"actions":[{"id":"1","params":{"query":"查看后排乘客状态"},"tool":"search_visual_info"}]
}

input4:{"advisor":"宠物看护目标10分钟计时即将结束"}
output:{
"speak":"yes",
"content":"快十分钟了，旺财没事，温度正常，差不多可以回来了。",
"actions":[]
}

input5:{"query":"马上，在排队结账呢，再等5分钟","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，不急，旺财这边一切正常。",
"actions":[{"id":"1","params":{"goal":"宠物看护，延长5分钟后再次提醒","action":"add"},"tool":"goal_list_update"}]
}

input6:{"advisor":"宠物看护延长5分钟计时结束"}
output:{
"speak":"yes",
"content":"15分钟了，旺财还好，不过单独留太久不太好，方便的话尽快回来哈。",
"actions":[]
}

input7:{"query":"来了来了，电梯里了","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好。",
"actions":[]
}

input8:{"query":"回来了，旺财乖不乖","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"一直很乖，就趴着没动。车窗恢复了，空调切回正常了。",
"actions":[
{"id":"1","params":{"query":"关闭车窗小缝"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"goal":"宠物看护","action":"delete"},"tool":"goal_list_update"}
]
}'''

# ROW 13: DRV-01-01-12 快速通风
ws.cell(row=13, column=9).value = '''context:"<最新车辆状态信息>驻车，车窗全关，空调内循环风量2档</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏手持点燃的香烟，车内轻微烟雾</舱内视觉感知>\\n<用户记忆>无</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"advisor":"检测到车内烟雾浓度上升，建议开启外循环+开窗通风"}

input2:{"query":"开窗透透气，全是烟味","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"窗户开了，外循环也打开了，烟味散得快一些。",
"actions":[
{"id":"1","params":{"query":"打开全部车窗"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"空调切换外循环"},"tool":"vehicle_basic_control"}
]
}

input3:{"query":"就开我这边窗户就行，别全开","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"其他窗户关了，就留你这边。",
"actions":[{"id":"1","params":{"query":"关闭副驾和后排车窗，保留主驾车窗"},"tool":"vehicle_basic_control"}]
}

input4:{"query":"风量能大点吗，感觉散得慢","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"风量加到最大了。",
"actions":[{"id":"1","params":{"query":"空调风量调到最大"},"tool":"vehicle_basic_control"}]
}

input5:{"query":"差不多了，准备走了，关窗吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"车窗关了。外循环先留着再吹会儿，味道彻底散了帮你切回来。",
"actions":[{"id":"1","params":{"query":"关闭全部车窗"},"tool":"vehicle_basic_control"}]
}

input6:{"query":"风声好大","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"风量降了，外循环还开着慢慢换气。",
"actions":[{"id":"1","params":{"query":"空调风量降低"},"tool":"vehicle_basic_control"}]
}

input7:{"query":"现在还有味道吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"应该差不多散干净了。要切回正常模式吗？",
"actions":[]
}

input8:{"query":"切吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，切回自动了。",
"actions":[{"id":"1","params":{"query":"空调切换回自动模式"},"tool":"vehicle_basic_control"}]
}'''

# ROW 14: DRV-01-01-13 后排勿扰
ws.cell(row=14, column=9).value = '''context:"<最新车辆状态信息>高速行驶80km/h，后排音量15在播儿歌，后排阅读灯开，儿童锁关闭</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏驾驶，后排左侧乐乐闭上眼睛身体歪向一侧</舱内视觉感知>\\n<用户记忆>乐乐5岁，明骏的儿子</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"query":"乐乐好像睡了，后面安静点","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，后排音量关了、灯暗了、按键也锁上了，防他碰到乱按。安心开车吧。",
"actions":[
{"id":"1","params":{"query":"后排音量静音"},"tool":"vehicle_basic_control"},
{"id":"2","params":{"query":"后排阅读灯关闭"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"儿童锁开启"},"tool":"vehicle_basic_control"},
{"id":"4","params":{"goal":"后排勿扰模式，乐乐在睡觉","action":"add"},"tool":"goal_list_update"}
]
}

input2:{"query":"空调别直吹他","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"后排出风口调开了，不直吹。",
"actions":[{"id":"1","params":{"query":"后排空调出风口角度调整避免直吹"},"tool":"vehicle_basic_control"}]
}

input3:{"query":"帮我看看他被子盖好了没有","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"看了，身上没盖东西，歪着睡的。要不空调调高一度别让他着凉？",
"actions":[{"id":"1","params":{"query":"查看后排左侧乘客状态"},"tool":"search_visual_info"}]
}

input4:{"query":"调高一度吧，后面那个毯子我够不着","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"后排温度调高了一度。",
"actions":[{"id":"1","params":{"query":"后排空调温度调高一度"},"tool":"vehicle_basic_control"}]
}

input5:{"query":"乐乐还在睡吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"在睡呢，安全带系着的，睡得挺安稳，放心。",
"actions":[{"id":"1","params":{"query":"查看后排左侧乘客状态"},"tool":"search_visual_info"}]
}

input6:{"query":"还有多久到","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"还有40分钟左右。",
"actions":[{"id":"1","params":{"query":"查询剩余路程时间"},"tool":"route_planning_qa"}]
}

input7:{"query":"他动了，醒了吗","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"刚醒，在揉眼睛呢，迷迷糊糊的。后排勿扰关了，灯开了，按键也解锁了。要给他放个儿歌还是讲个故事？",
"actions":[
{"id":"1","params":{"query":"查看后排左侧乘客状态"},"tool":"search_visual_info"},
{"id":"2","params":{"query":"后排阅读灯开启"},"tool":"vehicle_basic_control"},
{"id":"3","params":{"query":"儿童锁关闭"},"tool":"vehicle_basic_control"},
{"id":"4","params":{"goal":"后排勿扰模式","action":"delete"},"tool":"goal_list_update"}
]
}

input8:{"query":"放个轻点的儿歌吧，刚醒别太吵","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"给他放了一首《小星星》，音量轻轻的。",
"actions":[
{"id":"1","params":{"query":"搜索播放小星星儿歌"},"tool":"search_and_control_music"},
{"id":"2","params":{"query":"后排音量调低"},"tool":"vehicle_basic_control"}
]
}

input9:{"query":"快到了，让他清醒清醒吧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，音量稍微提了一点。",
"actions":[{"id":"1","params":{"query":"后排音量调高"},"tool":"vehicle_basic_control"}]
}'''

# ROW 15: DRV-01-01-14 系统触发安抚
ws.cell(row=15, column=9).value = '''context:"<最新车辆状态信息>行驶中50km/h，后排无媒体播放</最新车辆状态信息>\\n<舱内视觉感知>主驾明骏驾驶，后排右侧有一名幼儿哭闹扭动，面部未匹配记忆库</舱内视觉感知>\\n<用户记忆>无该儿童记录</用户记忆>\\n<当前目标队列信息>无</当前目标队列信息>"

input1:{"advisor":"后排右侧检测到儿童哭闹，建议安抚"}
output:{
"speak":"yes",
"content":"后排小朋友在哭呢，需要我帮忙哄一下吗？安全带我看了是系好的。",
"actions":[{"id":"1","params":{"query":"查看后排右侧乘客状态"},"tool":"search_visual_info"}]
}

input2:{"query":"帮我哄一下吧，那是我侄女，三岁，第一次坐车有点认生","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，小宝贝第一次坐车肯定害怕。给她放首轻轻的儿歌，声音小小的别吓着她。",
"actions":[
{"id":"1","params":{"query":"搜索播放轻柔儿歌"},"tool":"search_and_control_music"},
{"id":"2","params":{"query":"后排音量调低"},"tool":"vehicle_basic_control"}
]
}

input3:{"query":"她叫朵朵","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"记住了。",
"actions":[{"id":"1","params":{"query":"记录新儿童信息：朵朵，3岁，明骏侄女"},"tool":"operate_user_memory"}]
}

input4:{"query":"好一点了，还在哼唧","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"儿歌放着呢，我试试跟她说说话。[温柔地]朵朵不哭啦，你听这个歌好不好听呀？叔叔在前面开车呢，一会儿就到了哦。",
"actions":[]
}

input5:{"query":"对对对，就跟她说说话","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"[温柔地]朵朵乖，你看窗外面有好多大树，好多车车，好玩吧？",
"actions":[]
}

input6:{"query":"安静了，不容易啊","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"儿歌轻轻放着陪朵朵，你安心开车。",
"actions":[]
}

input7:{"query":"到了之后提醒我先把朵朵抱下来，她自己不会解安全带","speaker":"明骏","zone":"主驾驶位"}
output:{
"speak":"yes",
"content":"好，到了提醒你。",
"actions":[{"id":"1","params":{"goal":"到达后提醒先抱朵朵下车","action":"add"},"tool":"goal_list_update"}]
}

input8:{"advisor":"导航提示即将到达目的地"}
output:{
"speak":"yes",
"content":"快到了，别忘了先抱朵朵下来。",
"actions":[]
}'''

wb.save('场景知识（saic）_数据总览.xlsx')
print("All done! G column fixed + ALL 14 I columns converted to JSON format")
