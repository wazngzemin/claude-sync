# -*- coding: utf-8 -*-
"""
构建 IBS-M 专属每日健康追踪表（健康追踪表.xlsx）
两个 sheet：
  1. 每日记录 —— 针对慢性腹泻/IBS-M 的全方位逐日追踪
  2. 周总结   —— 自动算出每天综合状态分、找出最佳/最差日及原因
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

START_DATE = datetime.date(2026, 6, 19)   # 今天
N_DAYS = 30                                # 预填 30 天（覆盖 4 周）
HEADER_ROW = 1
DATA_START = 2                             # 数据从第 2 行开始
DATA_END = DATA_START + N_DAYS - 1

WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

# ---- 配色 ----
C_INTAKE   = "DCE6F1"  # 饮食/晨起 浅蓝
C_GUT      = "FCE4D6"  # 肠道症状 浅橙
C_SYMPTOM  = "FFF2CC"  # 身体症状 浅黄
C_WORK     = "E2EFDA"  # 精力/专注（工作）浅绿
C_MOOD     = "EAD1DC"  # 情绪/压力 浅紫
C_SLEEP    = "D9E1F2"  # 睡眠 浅靛
C_OTHER    = "F2F2F2"  # 其他 浅灰
C_SCORE    = "FFE699"  # 综合分 金
C_HEAD_TXT = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 列定义：(表头, 列宽, 分组色, 说明/提示)
COLS = [
    ("日期",            12, C_OTHER,   ""),
    ("星期",            7,  C_OTHER,   ""),
    ("晨起温水",        9,  C_INTAKE,  "是/否（起床先喝温水）"),
    ("早餐时间",        9,  C_INTAKE,  "如 7:30；没吃留空"),
    ("早餐内容",        22, C_INTAKE,  "吃了什么"),
    ("午餐时间",        9,  C_INTAKE,  ""),
    ("午餐内容",        22, C_INTAKE,  ""),
    ("晚餐时间",        9,  C_INTAKE,  ""),
    ("晚餐内容",        22, C_INTAKE,  ""),
    ("吃了凉/生冷",     11, C_INTAKE,  "是/否（冷饮/凉菜/生水果/凉水）"),
    ("触发食物记录",    20, C_INTAKE,  "怀疑哪样东西引起不适"),
    ("排便次数",        9,  C_GUT,     "整数"),
    ("排便性状",        12, C_GUT,     "干硬/正常成形/糊状/稀水样"),
    ("是否腹泻",        9,  C_GUT,     "是/否"),
    ("腹胀",            7,  C_SYMPTOM, "0无-5重"),
    ("排气",            7,  C_SYMPTOM, "0无-5重"),
    ("胃部凉感",        9,  C_SYMPTOM, "0无-5重"),
    ("疲劳",            7,  C_SYMPTOM, "0无-5重"),
    ("上午精力",        9,  C_WORK,    "1差-5好"),
    ("下午精力",        9,  C_WORK,    "1差-5好"),
    ("晚上精力",        9,  C_WORK,    "1差-5好"),
    ("上午专注",        9,  C_WORK,    "1差-5好"),
    ("下午专注",        9,  C_WORK,    "1差-5好"),
    ("晚上专注",        9,  C_WORK,    "1差-5好"),
    ("情绪",            7,  C_MOOD,    "1差-5好"),
    ("压力",            7,  C_MOOD,    "1低-5高"),
    ("入睡时间",        9,  C_SLEEP,   "如 23:30"),
    ("起床时间",        9,  C_SLEEP,   "如 7:00"),
    ("睡眠时长h",       9,  C_SLEEP,   "小时数，如 7.5"),
    ("睡眠质量",        9,  C_SLEEP,   "1差-5好"),
    ("运动(分钟)",      10, C_OTHER,   "如 30；没动填0"),
    ("饮水量ml",        9,  C_OTHER,   "全天温水总量"),
    ("用药/益生菌",     16, C_OTHER,   "中药/益生菌/其他"),
    ("备注",            26, C_OTHER,   "任何想记的"),
    ("综合状态分",      11, C_SCORE,   "自动计算(0-100)"),
]

# 列字母映射
def col_idx(name):
    for i, c in enumerate(COLS):
        if c[0] == name:
            return i + 1
    raise KeyError(name)

L = {c[0]: get_column_letter(i + 1) for i, c in enumerate(COLS)}

wb = Workbook()

# ============================================================
# Sheet 1: 每日记录
# ============================================================
ws = wb.active
ws.title = "每日记录"
ws.sheet_view.showGridLines = False

# 表头
for i, (name, width, color, note) in enumerate(COLS):
    cidx = i + 1
    ws.column_dimensions[get_column_letter(cidx)].width = width
    cell = ws.cell(row=HEADER_ROW, column=cidx, value=name)
    cell.font = Font(bold=True, size=10, color="3B3B3B")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    if note:
        # 批注作为提示写入第一数据行上方？用单元格 comment
        from openpyxl.comments import Comment
        cell.comment = Comment(note, "tracker")

ws.row_dimensions[HEADER_ROW].height = 38

# 综合分公式
def score_formula(r):
    # 仅当当天「是否腹泻」已填（说明这天确实记录了）才计算，否则留空，
    # 避免未来空白天被周总结误判为「最差日」。
    return (
        f'=IF($N{r}="","",ROUND('
        # 肠道 40 分：无腹泻 +20；三项肠道症状越低越好 (15-和)/15*20
        f'IF($N{r}="否",20,IF($N{r}="是",0,0))'
        f'+(15-($O{r}+$P{r}+$Q{r}))/15*20'
        # 精力+专注 25 分（六项满分 30）
        f'+($S{r}+$T{r}+$U{r}+$V{r}+$W{r}+$X{r})/30*25'
        # 情绪+压力 20 分（情绪高好、压力低好）
        f'+($Y{r}+(6-$Z{r}))/10*20'
        # 睡眠质量 15 分
        f'+$AD{r}/5*15'
        f',1))'
    )

# 数据行：预填日期 + 星期 + 综合分公式
for k in range(N_DAYS):
    r = DATA_START + k
    d = START_DATE + datetime.timedelta(days=k)
    ws.cell(row=r, column=col_idx("日期"), value=d.strftime("%Y-%m-%d"))
    ws.cell(row=r, column=col_idx("星期"), value=WEEKDAY_CN[d.weekday()])
    ws.cell(row=r, column=col_idx("综合状态分"), value=score_formula(r))
    # 周一加一条粗上边框分隔周
    is_week_start = d.weekday() == 0
    for i in range(len(COLS)):
        c = ws.cell(row=r, column=i + 1)
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.font = Font(size=10)
        # 轻微底纹区分分组（数据区淡化）
        if is_week_start:
            c.border = Border(left=thin, right=thin,
                              top=Side(style="medium", color="808080"), bottom=thin)
    # 周末日期标灰
    if d.weekday() >= 5:
        ws.cell(row=r, column=col_idx("日期")).fill = PatternFill("solid", fgColor="EFEFEF")
        ws.cell(row=r, column=col_idx("星期")).fill = PatternFill("solid", fgColor="EFEFEF")

# 冻结：前两列 + 表头
ws.freeze_panes = "C2"

# ---- 数据验证下拉 ----
def add_dv(formula, col_names, allow_blank=True):
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "请从下拉中选择"
    dv.errorTitle = "输入无效"
    ws.add_data_validation(dv)
    for name in col_names:
        c = L[name]
        dv.add(f"{c}{DATA_START}:{c}{DATA_END}")

add_dv('"是,否"', ["晨起温水", "吃了凉/生冷", "是否腹泻"])
add_dv('"干硬,正常成形,糊状,稀水样"', ["排便性状"])
add_dv('"0,1,2,3,4,5"', ["腹胀", "排气", "胃部凉感", "疲劳"])
add_dv('"1,2,3,4,5"', ["上午精力", "下午精力", "晚上精力",
                        "上午专注", "下午专注", "晚上专注",
                        "情绪", "压力", "睡眠质量"])

# 综合分 0-100 色阶（红-黄-绿）
score_col = L["综合状态分"]
ws.conditional_formatting.add(
    f"{score_col}{DATA_START}:{score_col}{DATA_END}",
    ColorScaleRule(
        start_type="num", start_value=40, start_color="F8696B",
        mid_type="num",   mid_value=65,  mid_color="FFEB84",
        end_type="num",   end_value=90,  end_color="63BE7B",
    ),
)

# ============================================================
# Sheet 2: 周总结
# ============================================================
ws2 = wb.create_sheet("周总结")
ws2.sheet_view.showGridLines = False
for col, w in {"A": 16, "B": 26, "C": 26, "D": 22, "E": 18, "F": 18}.items():
    ws2.column_dimensions[col].width = w

def title(r, text):
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws2.cell(row=r, column=1, value=text)
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[r].height = 26

def subhead(r, c, text):
    cell = ws2.cell(row=r, column=c, value=text)
    cell.font = Font(bold=True, size=10, color="3B3B3B")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

def kv(r, c, label, value):
    lc = ws2.cell(row=r, column=c, value=label)
    lc.font = Font(bold=True, size=10)
    lc.fill = PatternFill("solid", fgColor="F2F2F2")
    lc.border = border
    lc.alignment = Alignment(horizontal="right", vertical="center")
    vc = ws2.cell(row=r, column=c + 1, value=value)
    vc.border = border
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return vc

DREF = "每日记录"
srng = f"{DREF}!${score_col}${DATA_START}:${score_col}${DATA_END}"
drng = f"{DREF}!$A${DATA_START}:$A${DATA_END}"

r = 1
title(r, "📊 周总结 · 自动分析（填满数据后自动更新）")
r += 2

# ---- 区块一：最佳/最差日 ----
title(r, "① 哪天状态最好 / 最差，为什么")
r += 1
hdr = r
for c, t in enumerate(["", "🏆 最佳日", "⚠️ 最差日"], start=1):
    subhead(hdr, c, t)
r += 1

# 行项目：用 INDEX/MATCH 取 MAX/MIN 分数那天
rows_spec = [
    ("日期",        "日期"),
    ("综合状态分",  "综合状态分"),
    ("早餐内容",    "早餐内容"),
    ("午餐内容",    "午餐内容"),
    ("晚餐内容",    "晚餐内容"),
    ("吃了凉/生冷", "吃了凉/生冷"),
    ("是否腹泻",    "是否腹泻"),
    ("情绪(1-5)",   "情绪"),
    ("压力(1-5)",   "压力"),
    ("睡眠时长h",   "睡眠时长h"),
    ("睡眠质量",    "睡眠质量"),
    ("运动(分钟)",  "运动(分钟)"),
    ("备注",        "备注"),
]
# MAX/MIN 仅在有数值时
max_match = f'MATCH(MAX({srng}),{srng},0)'
min_match = f'MATCH(MIN({srng}),{srng},0)'
for label, field in rows_spec:
    col_letter = L[field]
    field_rng = f"{DREF}!${col_letter}${DATA_START}:${col_letter}${DATA_END}"
    lc = ws2.cell(row=r, column=1, value=label)
    lc.font = Font(bold=True, size=10)
    lc.fill = PatternFill("solid", fgColor="F2F2F2")
    lc.border = border
    lc.alignment = Alignment(horizontal="right", vertical="center")
    best = f'=IFERROR(INDEX({field_rng},{max_match}),"—")'
    worst = f'=IFERROR(INDEX({field_rng},{min_match}),"—")'
    bc = ws2.cell(row=r, column=2, value=best)
    wc = ws2.cell(row=r, column=3, value=worst)
    for cc in (bc, wc):
        cc.border = border
        cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    if field == "综合状态分":
        bc.font = Font(bold=True, color="2E7D32", size=11)
        wc.font = Font(bold=True, color="C62828", size=11)
    r += 1
r += 1

# ---- 区块二：本周平均/统计 ----
title(r, "② 本周关键指标统计")
r += 1
stat_specs = [
    ("记录天数",        f'=COUNT({srng})'),
    ("平均综合分",      f'=IFERROR(ROUND(AVERAGE({srng}),1),"—")'),
    ("腹泻天数",        f'=COUNTIF({DREF}!${L["是否腹泻"]}${DATA_START}:${L["是否腹泻"]}${DATA_END},"是")'),
    ("吃早餐天数",      f'=COUNTIF({DREF}!${L["早餐时间"]}${DATA_START}:${L["早餐时间"]}${DATA_END},"<>")'),
    ("吃凉/生冷天数",   f'=COUNTIF({DREF}!${L["吃了凉/生冷"]}${DATA_START}:${L["吃了凉/生冷"]}${DATA_END},"是")'),
    ("平均腹胀",        f'=IFERROR(ROUND(AVERAGE({DREF}!${L["腹胀"]}${DATA_START}:${L["腹胀"]}${DATA_END}),1),"—")'),
    ("平均压力",        f'=IFERROR(ROUND(AVERAGE({DREF}!${L["压力"]}${DATA_START}:${L["压力"]}${DATA_END}),1),"—")'),
    ("平均睡眠时长h",   f'=IFERROR(ROUND(AVERAGE({DREF}!${L["睡眠时长h"]}${DATA_START}:${L["睡眠时长h"]}${DATA_END}),1),"—")'),
    ("平均睡眠质量",    f'=IFERROR(ROUND(AVERAGE({DREF}!${L["睡眠质量"]}${DATA_START}:${L["睡眠质量"]}${DATA_END}),1),"—")'),
    ("平均饮水ml",      f'=IFERROR(ROUND(AVERAGE({DREF}!${L["饮水量ml"]}${DATA_START}:${L["饮水量ml"]}${DATA_END}),0),"—")'),
]
for i, (label, formula) in enumerate(stat_specs):
    col = 1 if i % 2 == 0 else 4
    kv(r if i % 2 == 0 else r, col, label, None)
    # set formula
    ws2.cell(row=r, column=(2 if i % 2 == 0 else 5)).value = formula
    if i % 2 == 1:
        r += 1
if len(stat_specs) % 2 == 1:
    r += 1
r += 1

# ---- 区块三：规律对照（自动算腹泻关联）----
title(r, "③ 自动规律发现（吃早餐 vs 吃凉 与腹泻的关系）")
r += 1
bk = L["早餐时间"]; cold = L["吃了凉/生冷"]; dia = L["是否腹泻"]
bk_rng = f'{DREF}!${bk}${DATA_START}:${bk}${DATA_END}'
cold_rng = f'{DREF}!${cold}${DATA_START}:${cold}${DATA_END}'
dia_rng = f'{DREF}!${dia}${DATA_START}:${dia}${DATA_END}'
rule_specs = [
    ("没吃早餐的天数里 · 腹泻率",
     f'=IFERROR(TEXT(COUNTIFS({bk_rng},"",{dia_rng},"是")/COUNTIFS({bk_rng},"",{dia_rng},"<>"),"0%"),"数据不足")'),
    ("吃了早餐的天数里 · 腹泻率",
     f'=IFERROR(TEXT(COUNTIFS({bk_rng},"<>",{dia_rng},"是")/COUNTIFS({bk_rng},"<>",{dia_rng},"<>"),"0%"),"数据不足")'),
    ("吃了凉/生冷的天数里 · 腹泻率",
     f'=IFERROR(TEXT(COUNTIFS({cold_rng},"是",{dia_rng},"是")/COUNTIFS({cold_rng},"是",{dia_rng},"<>"),"0%"),"数据不足")'),
    ("没吃凉的天数里 · 腹泻率",
     f'=IFERROR(TEXT(COUNTIFS({cold_rng},"否",{dia_rng},"是")/COUNTIFS({cold_rng},"否",{dia_rng},"<>"),"0%"),"数据不足")'),
]
for label, formula in rule_specs:
    vc = kv(r, 1, label, None)
    ws2.cell(row=r, column=2).value = formula
    ws2.cell(row=r, column=2).font = Font(bold=True, size=11, color="1565C0")
    r += 1
r += 1

# ---- 使用说明 ----
title(r, "④ 怎么用这张表")
r += 1
tips = [
    "每天晚上 10:30 我会主动找你，把当天数据念给你 / 问你，逐格填进「每日记录」。",
    "症状项 0=没有、5=很重；精力/专注/情绪/睡眠质量 1=最差、5=最好；压力 1=很松、5=很大。",
    "「综合状态分」自动算出来：肠道40分 + 精力专注25分 + 情绪压力20分 + 睡眠15分。",
    "排便性状参考布里斯托：干硬=便秘端，正常成形=理想，糊状/稀水样=腹泻端。",
    "记满一周后，本页自动找出你状态最好/最差的那天，并对照那天吃了什么、睡得怎样、压力多大。",
    "③ 区块会自动算出『没吃早餐 vs 吃了早餐』『吃凉 vs 没吃凉』各自的腹泻率——这是验证医生判断的关键证据。",
    "把这张表带去看消化科，比口头描述精确得多。",
]
for t in tips:
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws2.cell(row=r, column=1, value="• " + t)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    c.font = Font(size=10)
    ws2.row_dimensions[r].height = 30
    r += 1

wb.save("健康追踪表.xlsx")
print("OK saved 健康追踪表.xlsx")
print(f"每日记录: {N_DAYS} 天 ({START_DATE} 起), {len(COLS)} 列")
