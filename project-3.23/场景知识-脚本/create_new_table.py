from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "执行逻辑"

header_font = Font(bold=True, size=11, name="Arial")
header_fill = PatternFill('solid', start_color='D9E1F2')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='top', wrap_text=True)
cell_font = Font(size=10, name="Arial")

headers = ["序号", "场景名称", "执行逻辑（含JSON示例）"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 120

rows_data = []

# ========== ROW 1: 儿童娱乐·陪伴安抚 ==========
logic_1 = """1. 获取信息：
从用户语音中提取孩子称呼和位置，例如"后排的宝宝""副驾的小朋友""乐乐""我儿子""女儿"等。
如果用户记忆中有孩子信息，可以参考孩子的年龄、性别、娱乐偏好、聊天偏好，例如"乐乐喜欢奥特曼""小雨喜欢儿歌和猜谜"。
如果缺少关键信息，只补问一句，例如："宝宝是在后排吗？我给他放点故事还是玩个语音小游戏？"或"他大概几岁？我好选更合适的内容。"

2. 判断可用娱乐方式：
如果孩子在前排，优先推荐语音游戏、儿歌、故事、猜谜、成语接龙等不依赖屏幕的内容。
如果孩子在后排且有后排屏，可以推荐动画、儿童短视频、互动游戏等视觉内容。
如果孩子在后排但没有屏幕，优先推荐音频故事、儿歌、语音互动游戏。
如果位置不明确，不要假设屏幕能力，优先推荐通用的语音或音频内容。

3. 生成内容推荐：
根据孩子年龄和偏好组合，给出1-2个具体选项。
例如："我给他放一段奥特曼故事，或者来个恐龙猜谜，你想选哪个？"
例如："如果他现在精神挺好，可以玩个简单的猜谜游戏，我来出题。"

4. 向用户或孩子确认后执行：
不要一次性抛出太多选择。
如果用户已经明确要某个内容，比如"放奥特曼"，可以直接执行或给出最接近的内容。
如果涉及视频、屏幕、较大音量，尽量简短确认，避免影响车内其他人。

5. 结束或切换：
如果孩子已经安静下来，可以减少语音打扰，保持内容轻量播放。"""

examples_1 = """示例1：
输入：{"query": "给我儿子放点东西看", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：后排有车载屏幕可用
memory：儿子5岁，喜欢恐龙
输出：{"speak": "yes", "content": "[温柔]宝贝喜欢恐龙对吧，我用后排屏给他放一集益智恐龙动画可以吗？", "actions": [{"id": "play_dino_video", "params": {"keyword": "恐龙益智动画", "target": "rear_screen"}, "tool": "search_and_control_short_video"}]}

示例2：
输入：{"query": "帮我哄一下后面那个小祖宗", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：后排无屏幕
memory：无孩子相关记忆
输出：{"speak": "yes", "content": "[温柔]宝宝大概几岁呀，我好给他选合适的安抚内容", "actions": []}
（追问年龄后得知3岁）
输入：{"query": "3岁", "speaker": "主驾驶位", "zone": "主驾驶位"}
输出：{"speak": "yes", "content": "[温柔]那我给他放一首轻柔儿歌，再玩个简单猜小动物的语音小游戏，帮他安静下来好不好？", "actions": [{"id": "play_lullaby", "params": {"keyword": "低幼舒缓儿歌"}, "tool": "search_and_control_music"}]}

示例3：
输入：{"query": "孩子闹着太无聊了", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：孩子在副驾前排，前排无法播放视频
输出：{"speak": "yes", "content": "[开心]副驾不能看屏幕哦，我来给他出几个简单小谜语，轻松玩一会安静下来可以吗？", "actions": []}"""

rows_data.append(("儿童娱乐·陪伴安抚", logic_1 + "\n\n" + examples_1))

# ========== ROW 2: 乘客娱乐·后排观影 ==========
logic_2 = """1. 获取信息：
用户话语中表达的娱乐需求（如"想看电影""后排屏幕能看吗"）、位置信息（如"我在后排"）。
查询用户记忆中的观影历史偏好、聊天中提及的感兴趣内容类型。
2. 判断可用性：
考虑前排座位空间、屏幕设置等场景因素，评估当前前排观影的可行性。
3. 生成处理方案：根据用户可能的偏好（影片类型、画质、音效等），提供1-2个具体的前排观影方案，并向用户确认后执行。
4. 优化观影体验：结合前排场景要点，确保屏幕设置、音效等符合前排观影需求，提升用户观影放松效果。"""

examples_2 = """示例1：
输入：{"query": "后排屏幕能看电影吗", "speaker": "二排左", "zone": "二排左"}
env_info：后排吸顶屏可用
memory：无观影偏好记录
输出：{"speak": "yes", "content": "[开心]可以的，你想看什么类型的电影，我帮你推荐？", "actions": []}
（用户选择悬疑片后）
输入：{"query": "想看悬疑片", "speaker": "二排左", "zone": "二排左"}
输出：{"speak": "yes", "content": "[开心]给你推荐一部高分悬疑片，直接播放可以吗？", "actions": [{"id": "play_suspense", "params": {"keyword": "高分悬疑电影", "target": "rear_ceiling_screen"}, "tool": "search_and_control_short_video"}]}

示例2：
输入：{"query": "帮我找点片子打发下路上时间", "speaker": "二排右", "zone": "二排右"}
env_info：后排吸顶屏可用
memory：用户偏爱科幻题材
输出：{"speak": "yes", "content": "[开心]你平时喜欢科幻类，我给你选一部口碑不错的科幻大片，放后排屏刚好合适，要不要直接播放？", "actions": [{"id": "play_scifi", "params": {"keyword": "高分科幻电影", "target": "rear_ceiling_screen"}, "tool": "search_and_control_short_video"}]}

示例3：
输入：{"query": "关掉后排屏幕，不看了", "speaker": "二排左", "zone": "二排左"}
输出：{"speak": "yes", "content": "好的", "actions": [{"id": "close_rear_screen", "params": {"function": "rear_screen", "action": "off"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("乘客娱乐·后排观影", logic_2 + "\n\n" + examples_2))

# ========== ROW 3: 乘客娱乐·前排观影（驻车） ==========
logic_3 = """1. 获取信息：
用户话语中关于驻车时娱乐或观看影片的明确表达（如"驻车想看电影"）。
确认用户观影需求针对前排（如"前排看电影"）。
2. 判断前排观影条件：
考虑前排座位空间、屏幕设置等场景因素，评估当前前排观影的可行性。
3. 生成处理方案：根据用户可能的偏好（影片类型、画质、音效等），提供1-2个具体的前排观影方案，并向用户确认后执行。
4. 优化观影体验：结合前排场景要点，确保屏幕设置、音效等符合前排观影需求，提升用户观影放松效果。"""

examples_3 = """示例1：
输入：{"query": "等人无聊，放部电影打发时间", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车辆正常驻车，用户在主驾
输出：{"speak": "yes", "content": "[开心]好嘞，驻车状态可以安心观影，我帮你微调座椅、拉上遮阳帘，你想看什么类型影片？", "actions": [{"id": "seat_adjust", "params": {"function": "seat_recline", "zone": "主驾驶位", "angle": -10}, "tool": "vehicle_basic_control"}, {"id": "close_sunshade", "params": {"function": "sunshade", "action": "close"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "前排能看视频吗", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车辆处于行驶状态（D挡）
输出：{"speak": "yes", "content": "[关心]为了行车安全，行驶中前排不能播放视频哦，等车辆停稳驻车后再看，我先给你放点舒缓音乐可以吗？", "actions": [{"id": "play_relax_music", "params": {"keyword": "舒缓轻音乐"}, "tool": "search_and_control_music"}]}

示例3：
输入：{"query": "充电还要好久，给我看点剧", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车辆驻车充电中
memory：上次观看剧集未播完
输出：{"speak": "yes", "content": "[开心]充电还有一段时间，刚好可以追剧，要不要接着播放你上次没看完的剧集？", "actions": [{"id": "resume_show", "params": {"action": "resume_last", "target": "front_screen"}, "tool": "search_and_control_short_video"}, {"id": "set_volume", "params": {"function": "volume", "level": "comfortable"}, "tool": "vehicle_basic_control"}, {"id": "ambient_light", "params": {"function": "ambient_light", "brightness": "low", "color": "warm"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("乘客娱乐·前排观影（驻车）", logic_3 + "\n\n" + examples_3))

# ========== ROW 4: 车内环境·急速降温 ==========
logic_4 = """1. 启动降温流程：
识别车门开启、用户上车行为信号，同步检测车内环境温度，当温度超出舒适阈值时，自动触发全车极速降温联动策略。
自动将空调调至最大制冷风量档位，按需开启主副驾座椅通风，全车车窗、天窗一次性拉满最大开启幅度，强化内外空气对流散热。
2. 动态调整车窗状态：
行车过程中实时监测车速，根据车速区间调整车窗和天窗的打开百分比（如低速时保持较大开启度，中高速时逐步减小开启度至关闭），以平衡降温效果与风噪控制。车辆行驶中实时采集实时车速数据，按低速、中速、高速不同车速区间，分级阶梯式自动调整车窗升降开度百分比、天窗开启角度。
低速区间保持较大开窗幅度保证通风；中高速区间逐步收窄开窗开度、适时关闭天窗，从源头减少气流灌入，实现通风和风噪的智能平衡。
3. 优化降温体验：
全程实时监测车内温度回落变化，待车内温度降至人体舒适区间后，自动循序关闭全部车窗与天窗。
空调由最大制冷模式切换为日常恒温正常模式，保留座椅通风按需运行，维持长久舒适乘坐环境，无多余操作打扰。"""

examples_4 = """示例1：
输入：{"query": "热死了，快降降温", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车内温度43℃，车辆驻车未行驶
输出：{"speak": "yes", "content": "[关心]车内温度偏高，马上开启全车极速降温模式", "actions": [{"id": "ac_max_cool", "params": {"function": "air_conditioning", "mode": "max_cooling", "circulation": "internal"}, "tool": "vehicle_basic_control"}, {"id": "seat_vent_on", "params": {"function": "seat_ventilation", "zone": "all_front", "level": 3}, "tool": "vehicle_basic_control"}, {"id": "window_all_open", "params": {"function": "window", "position": "all", "action": "open", "percentage": 100}, "tool": "vehicle_basic_control"}, {"id": "sunroof_open", "params": {"function": "sunroof", "action": "open", "percentage": 100}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "风噪太大了，把窗户关小一点", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车速85km/h，降温流程进行中
输出：{"speak": "yes", "content": "好的，马上调整开窗幅度降低风噪", "actions": [{"id": "window_reduce", "params": {"function": "window", "position": "all", "action": "set", "percentage": 10}, "tool": "vehicle_basic_control"}, {"id": "sunroof_close", "params": {"function": "sunroof", "action": "close"}, "tool": "vehicle_basic_control"}]}

示例3：
输入：{"advisor": "车内温度已降至26℃，达到舒适区间"}
env_info：车辆中速行驶中，降温流程执行中
输出：{"speak": "yes", "content": "[开心]车内已达到舒适温度，帮你关闭车窗，切换正常制冷模式", "actions": [{"id": "window_close_all", "params": {"function": "window", "position": "all", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "sunroof_close", "params": {"function": "sunroof", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "ac_normal", "params": {"function": "air_conditioning", "mode": "auto", "temperature": 24}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("车内环境·急速降温", logic_4 + "\n\n" + examples_4))

# ========== ROW 5: 车内环境·急速升温 ==========
logic_5 = """1. 启动核心功能：
开启空调制热模式并调至合适温度，同时开启座椅加热和方向盘加热。通过语音、场景自动、手动三类触发源启动核心功能，遵循接触式体感优先原则，按车内初始温度分级适配空调制热、座椅及方向盘加热参数，同步做好冷量隔绝、内循环安全管控与共情式语音交互。
2. 协同作用调节：
综合运用空调、座椅加热、方向盘加热等多种功能协同作用，关注不同功能的启动顺序和强度调节。统筹空调、加热、风道、除雾多系统联动，按时序、场景诉求、行车安全、分区能耗、强度梯度实现精细化协同调节。
3. 动态监控调整：
在升温过程中，关注车内温度变化，适时调整各功能参数，确保在满足快速升温的同时避免温度过高或能源浪费。全周期多维度监控车内核心数据，按温度节点分阶段自动调整参数、适配异常场景，同时做好安全兜底与用户习惯学习，在保障快速升温与舒适体感的同时规避温度过高与能源浪费。"""

examples_5 = """示例1：
输入：{"query": "冻死了", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车内温度2℃
输出：{"speak": "yes", "content": "[关心]这也太冷了，暖气全开了", "actions": [{"id": "ac_heat", "params": {"function": "air_conditioning", "mode": "heating", "temperature": 28}, "tool": "vehicle_basic_control"}, {"id": "seat_heat_on", "params": {"function": "seat_heating", "zone": "主驾驶位", "level": 3}, "tool": "vehicle_basic_control"}, {"id": "steering_heat", "params": {"function": "steering_wheel_heating", "action": "on"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "手好冰", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车内温度8℃
输出：{"speak": "yes", "content": "[关心]方向盘加热给你开了，暖暖手", "actions": [{"id": "steering_heat", "params": {"function": "steering_wheel_heating", "action": "on"}, "tool": "vehicle_basic_control"}]}

示例3：
输入：{"advisor": "车内温度已升至22℃"}
env_info：升温流程执行中，座椅加热3档运行
输出：{"speak": "yes", "content": "[开心]暖和多了吧，座椅加热我给你降到1档", "actions": [{"id": "seat_heat_down", "params": {"function": "seat_heating", "zone": "主驾驶位", "level": 1}, "tool": "vehicle_basic_control"}, {"id": "ac_adjust", "params": {"function": "air_conditioning", "mode": "heating", "temperature": 24}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("车内环境·急速升温", logic_5 + "\n\n" + examples_5))

# ========== ROW 6: 座椅调节·副驾座椅调节 ==========
logic_6 = """1. 获取信息：
用户话语中关于座椅舒适的诉求或抱怨（如"座椅有点硬""靠背不舒服"）。
2. 确定基础调节项：
关注座椅的靠背角度、坐垫高度和脚托等基础调节项，作为提升舒适性的关键物理调节。将座椅全维度物理调节作为舒适性提升的核心基础，包括：靠背倾斜角度、坐垫整体高度、坐垫前后位置、坐垫前端倾角、腰部支撑（凸起度 + 上下位置）、头枕高度与前后角度、腿托伸缩长度。
3. 考虑功能调节：
根据用户实际需求，考虑按摩、通风、加热等功能调节，选择合适的功能来增强舒适感。在基础物理调节的基础上，根据用户诉求、环境条件与行驶场景，主动匹配对应的舒适功能，实现"姿态舒适 + 功能舒适"双重提升。
4. 综合调节：
综合运用各项调节项，以全面提升用户对座椅舒适性的满意度。针对用户的综合诉求，生成"基础物理调节 + 功能调节"的一体化组合方案，而非单一调节。例如用户说"开长途腰酸背痛"，自动生成"腰部支撑上调 1 档 + 靠背向后调 3°+ 开启腰部按摩 2 档"的组合方案；用户说"今天好热，坐着不舒服"，自动生成"坐垫调低 1 档 + 开启座椅通风 2 档"的组合方案。"""

examples_6 = """示例1：
输入：{"query": "靠背不太舒服", "speaker": "副驾驶位", "zone": "副驾驶位"}
输出：{"speak": "yes", "content": "[温柔]我给你调一下，往后靠一点试试", "actions": [{"id": "recline_adjust", "params": {"function": "seat_recline", "zone": "副驾驶位", "angle": -5}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "坐久了腰酸", "speaker": "副驾驶位", "zone": "副驾驶位"}
输出：{"speak": "yes", "content": "[关心]按摩给你开着，腰部重点按一下", "actions": [{"id": "massage_on", "params": {"function": "seat_massage", "zone": "副驾驶位", "level": 2, "mode": "lumbar"}, "tool": "vehicle_basic_control"}]}

示例3：
输入：{"query": "想躺一会", "speaker": "副驾驶位", "zone": "副驾驶位"}
env_info：车辆驻车状态
输出：{"speak": "yes", "content": "[温柔]好，我给你放到最舒服的角度，脚托也伸出来", "actions": [{"id": "recline_flat", "params": {"function": "seat_recline", "zone": "副驾驶位", "angle": -40}, "tool": "vehicle_basic_control"}, {"id": "footrest_out", "params": {"function": "footrest", "zone": "副驾驶位", "action": "extend"}, "tool": "vehicle_basic_control"}, {"id": "massage_gentle", "params": {"function": "seat_massage", "zone": "副驾驶位", "level": 1}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("座椅调节·副驾座椅调节", logic_6 + "\n\n" + examples_6))

# ========== ROW 7: 乘客娱乐·5D影院 ==========
logic_7 = """1. 获取信息：
影片声音信号和画面内容。实时采集影片音频声音信号与视频画面内容，同步解析剧情节奏、场景氛围、音效强弱、画面色调及动态变化等核心信息，为后续智能调控提供数据依据。
2. 控制多系统协调运作：
联动座椅振子、氛围灯、香氛系统，调整影院音效，关闭车窗和遮阳帘。统一调度整车多模块联动运行，包含：座椅震动振子、车内氛围灯、智能香氛系统、影院级音效系统，同时自动控制车窗升降与遮阳帘开合，实现多设备一体化协同管控。
3. 精准匹配场景：
结合影片声音信号和画面内容精准控制各系统，以匹配不同的观影场景。基于解析后的影片音频、画面内容，对各子系统进行精细化参数调节，针对动作、悬疑、温情、惊悚等不同观影场景，匹配对应的震动幅度、灯光色彩、香氛香型及音效模式，做到场景与环境氛围高度契合。
4. 确保效果同步：
在观影过程中，确保各系统响应及时、效果同步，从而强化沉浸式的观影感受。观影全程实时同步各系统响应时序，严格把控指令延迟，确保座椅震动、氛围灯变化、香氛释放、音效调校与影片画面音效毫秒级同步；全程动态自适应调节，无滞后、无错位，最大化强化 5D 沉浸式观影体验。"""

examples_7 = """示例1：
输入：{"query": "看个5D电影", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车辆驻车中
输出：{"speak": "yes", "content": "[开心]好，车窗遮阳帘我都给你关上，音效切影院模式了，准备好了吗？", "actions": [{"id": "close_window", "params": {"function": "window", "position": "all", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "close_sunshade", "params": {"function": "sunshade", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "cinema_audio", "params": {"function": "audio_mode", "mode": "cinema"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"advisor": "影片播放中检测到爆炸场景"}
env_info：5D影院模式运行中
输出：{"speak": "no", "content": "", "actions": [{"id": "vibrate_strong", "params": {"function": "seat_vibrator", "intensity": "strong"}, "tool": "vehicle_basic_control"}, {"id": "light_red", "params": {"function": "ambient_light", "color": "red", "mode": "flash"}, "tool": "vehicle_basic_control"}, {"id": "bass_boost", "params": {"function": "audio_equalizer", "bass": "enhanced"}, "tool": "vehicle_basic_control"}]}

示例3：
输入：{"query": "震得有点猛", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：5D影院模式运行中
输出：{"speak": "yes", "content": "好，振动给你调小点", "actions": [{"id": "vibrate_reduce", "params": {"function": "seat_vibrator", "intensity": "medium"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("乘客娱乐·5D影院", logic_7 + "\n\n" + examples_7))

# ========== ROW 8: 减压放松·冥想空间 ==========
logic_8 = """1. 获取信息：
用户话语中关于压力大、身体疲惫、想要放松的表述。精准捕捉用户语音表述中压力过大、身心疲惫、情绪紧绷、想要舒缓放松等相关语义，自动识别放松、减压、疗愈、舒缓情绪等潜在诉求。
2. 判断用户状态：
理解用户处于身心紧张状态，明确其对放松环境和引导的直接需求。结合用户表述，智能判断用户当前处于精神紧张、身心疲惫、情绪焦虑的状态，精准识别用户对专属放松环境、专业冥想引导、情绪舒缓疗愈的核心需求。
3. 提供冥想空间：
包含多种使用场景，以适应不同用户在不同情境下的放松需求。提供覆盖多时段、多情境的专属冥想空间模式，适配通勤驻车、午休小憩、睡前疗愈、工作减压、情绪平复等不同使用场景，满足用户随时随地的放松诉求。
4. 提供引导方式：
提供多种冥想引导方式，满足用户多样化的放松偏好，帮助其更好地进入放松状态。配备多种类型冥想引导形式，涵盖呼吸引导、身体渐进放松、正念冥想、自然声疗愈、助眠解压等不同风格，适配不同用户个人喜好，帮助用户快速平复情绪、卸下疲惫，平稳进入深度放松状态。"""

examples_8 = """示例1：
输入：{"query": "好累啊想放松一下", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车辆驻车中
输出：{"speak": "yes", "content": "[轻柔]我帮你进入冥想模式，先把座椅给你放平一点，灯光调暗，放段引导音乐，跟着我的节奏呼吸就好", "actions": [{"id": "seat_recline", "params": {"function": "seat_recline", "zone": "主驾驶位", "angle": -20}, "tool": "vehicle_basic_control"}, {"id": "light_dim", "params": {"function": "ambient_light", "brightness": "low", "color": "warm"}, "tool": "vehicle_basic_control"}, {"id": "close_window", "params": {"function": "window", "position": "all", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "play_meditation", "params": {"keyword": "冥想引导音频"}, "tool": "search_and_control_music"}, {"id": "massage_on", "params": {"function": "seat_massage", "zone": "主驾驶位", "level": 1}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "压力好大", "speaker": "主驾驶位", "zone": "主驾驶位"}
输出：{"speak": "yes", "content": "[轻柔]那静一静，我带你做个5分钟的呼吸练习，把其他东西都关掉", "actions": [{"id": "pause_media", "params": {"function": "media", "action": "pause"}, "tool": "vehicle_basic_control"}, {"id": "close_window", "params": {"function": "window", "position": "all", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "light_dim", "params": {"function": "ambient_light", "brightness": "low"}, "tool": "vehicle_basic_control"}, {"id": "breathing_guide", "params": {"keyword": "呼吸引导练习"}, "tool": "search_and_control_music"}]}

示例3：
输入：{"query": "好了差不多了", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：冥想模式已进行3分钟
输出：{"speak": "yes", "content": "[温柔]好，灯光慢慢调亮，座椅给你恢复", "actions": [{"id": "light_restore", "params": {"function": "ambient_light", "brightness": "normal"}, "tool": "vehicle_basic_control"}, {"id": "seat_restore", "params": {"function": "seat_recline", "zone": "主驾驶位", "angle": 0}, "tool": "vehicle_basic_control"}, {"id": "massage_off", "params": {"function": "seat_massage", "zone": "主驾驶位", "action": "off"}, "tool": "vehicle_basic_control"}, {"id": "stop_meditation", "params": {"action": "stop"}, "tool": "search_and_control_music"}]}"""

rows_data.append(("减压放松·冥想空间", logic_8 + "\n\n" + examples_8))

# ========== ROW 9: 短暂休憩·车内小憩 ==========
logic_9 = """1. 获取信息：
用户触发驻车休息需求（如"我想休息一会儿"）。
2. 调整休息环境：
调整座椅角度至适合休息的姿势，关闭车窗和遮阳帘以营造私密昏暗环境。
3. 维持适宜条件：
打开空调维持适宜温度，查询用户历史休息时长或主动询问是否需要叫醒服务。作为推荐时长。
4. 保障休息过程：
在休息过程中，持续确保环境的舒适性和安全性，满足用户短暂休息的核心需求。
5. 明确核心目标：
精准识别用户车内驻车小憩、短暂休息的核心诉求，整合座椅、车窗、遮阳帘、空调一体化环境调节，同时智能匹配叫醒服务、免打扰管控与温柔渐进唤醒，保障休息舒适度、私密性与全程安全。
6. 调节车内环境：
自动将座椅调整至休憩舒适角度，关闭车窗隔绝外界噪音，闭合遮阳帘营造昏暗私密休息氛围，打造安静避光的驻车休息环境。
7. 维持适宜条件：
开启空调恒温至舒适体感模式，调取用户历史休息时长作为默认参考，未明确时长则主动问询，同步预设叫醒服务时间。
8. 全程保障管控：
休息期间持续维持车内恒温与静谧环境，来电及普通通知静音静默仅屏幕显示、不语音播报；同步守护车辆安全，不打扰用户浅休息状态。
9. 梯度温柔唤醒：
到达设定时间后，氛围灯渐亮、轻柔语音提示、座椅缓慢复位，采用渐进式唤醒方式，避免突然发声惊扰用户。
10. 无指令默认适配：
用户未说明休息时长时，调取历史常用小憩时长静默设定闹钟，不主动打扰，支持随时语音修改时长。
11. 休憩后复位适配：
唤醒结束后，可语音问询是否将座椅、车窗、遮阳帘、空调恢复常规行车状态，便捷衔接后续行程。"""

examples_9 = """示例1：
输入：{"query": "我想眯一会", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车辆驻车中
输出：{"speak": "yes", "content": "[轻柔]好，座椅给你放平，窗帘拉上，空调调到舒服的温度，你想睡多久？", "actions": [{"id": "seat_flat", "params": {"function": "seat_recline", "zone": "主驾驶位", "angle": -35}, "tool": "vehicle_basic_control"}, {"id": "close_sunshade", "params": {"function": "sunshade", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "close_window", "params": {"function": "window", "position": "all", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "ac_quiet", "params": {"function": "air_conditioning", "mode": "quiet", "temperature": 24}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "20分钟叫我", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：小憩模式已启动
输出：{"speak": "yes", "content": "[轻柔]好，20分钟后叫你", "actions": [{"id": "set_alarm", "params": {"action": "add", "goal": "20分钟后温柔唤醒用户", "duration": 1200}, "tool": "goal_list_update"}, {"id": "mute_notifications", "params": {"function": "notification", "mode": "silent_display_only"}, "tool": "vehicle_basic_control"}]}

示例3：
输入：{"advisor": "小憩闹钟20分钟已到"}
env_info：用户小憩中，车内灯光关闭
输出：{"speak": "yes", "content": "[轻柔]到时间了，该起来了", "actions": [{"id": "light_gradual", "params": {"function": "ambient_light", "brightness": "gradual_increase"}, "tool": "vehicle_basic_control"}, {"id": "seat_restore", "params": {"function": "seat_recline", "zone": "主驾驶位", "angle": 0, "speed": "slow"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("短暂休憩·车内小憩", logic_9 + "\n\n" + examples_9))

# ========== ROW 10: 行车关怀·雨天关怀 ==========
logic_10 = """1. 监测触发条件：
系统实时监测车辆状态（是否处于行车过程中或用户切换至D挡）和天气状况（是否存在下雨、起雾、下雪等影响能见度的恶劣天气）。
2. 判断条件是否同时满足：
当车辆状态为行车中或切换至D挡，且天气状况为恶劣天气时，触发处理方案；否则不触发。
3. 执行安全保障操作：
自动开启前后雾灯、自动雨刮、后视镜及后挡风玻璃加热功能。
4. 发出语音提醒：
在执行自动操作的同时，语音提醒用户"当前天气恶劣，已为您开启相关安全设置，请注意行车安全"。"""

examples_10 = """示例1：
输入：{"advisor": "检测到降大雨，车辆行驶中"}
env_info：车辆行驶中，开始下大雨
输出：{"speak": "yes", "content": "[关心]下雨了，雨刮和雾灯我给你开了，注意路滑", "actions": [{"id": "wiper_on", "params": {"function": "wiper", "action": "auto"}, "tool": "vehicle_basic_control"}, {"id": "fog_light_front", "params": {"function": "fog_light", "position": "front", "action": "on"}, "tool": "vehicle_basic_control"}, {"id": "mirror_heat", "params": {"function": "mirror_heating", "action": "on"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"advisor": "检测到降雪，用户切换至D挡准备出发"}
env_info：室外下雪，用户刚挂D挡
输出：{"speak": "yes", "content": "[关心]外面在下雪，安全设置都给你开好了，开慢点", "actions": [{"id": "wiper_low", "params": {"function": "wiper", "action": "low_speed"}, "tool": "vehicle_basic_control"}, {"id": "fog_light_all", "params": {"function": "fog_light", "position": "all", "action": "on"}, "tool": "vehicle_basic_control"}, {"id": "mirror_heat", "params": {"function": "mirror_heating", "action": "on"}, "tool": "vehicle_basic_control"}, {"id": "rear_defog", "params": {"function": "rear_windshield_heating", "action": "on"}, "tool": "vehicle_basic_control"}]}

示例3：
输入：{"advisor": "雨已停止，路面干燥"}
env_info：雨刮和雾灯运行中
输出：{"speak": "no", "content": "", "actions": [{"id": "wiper_off", "params": {"function": "wiper", "action": "off"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("行车关怀·雨天关怀", logic_10 + "\n\n" + examples_10))

# ========== ROW 11: 车内关怀·宠物关怀 ==========
logic_11 = """1. 获取信息：
用户临时离开的时长（如"去买瓶水，很快回来"）。
查询宠物的品种、年龄、健康状况及是否有分离焦虑（如"我家金毛今年3岁，有点怕热"）。或基于舱内传感器检测宠物当前状态（是否安静、有无异常活动）。
2. 判断环境调节方案：
若短时离开（如5分钟内）→ 调节空调至适宜温度（夏季24-26℃，冬季18-22℃），可适当开启1-2指宽车窗缝隙通风。
若较长时间离开（如超过15分钟）→ 确保空调持续运行，关闭车窗（避免宠物跳出或外界干扰），检查车内是否有足够空间让宠物活动。
3. 提醒准备宠物所需物品：
若用户离开时长超过30分钟 →提醒预留适量清水（使用不易打翻的容器），若离开时间更长（如1小时以上），提醒补充少量宠物零食。
4. 确认安全状态：
离开前检查车内无危险物品（如尖锐物品、小零件），锁车后通过车机APP实时监测车内温度和宠物动态，确保宠物安全。
5. 场景意图识别：识别临时留车、行车躁动、长途自驾、途经复杂路段等各类携宠出行场景。
6. 全域状态感知：通过车内传感、音视频监测宠物情绪行为，同步采集温湿度、噪音、光照等环境数据。
7. 智能策略适配：自动匹配空调温控、车窗通风、安全锁、舒缓音乐等适配方案。
8. 安全风险干预：对高温密闭、宠物扒窗、独处焦虑等隐患主动预警并限制危险行为。
9. 实时值守告警：全程动态监测宠物与车内环境，异常情况及时向车主APP推送提醒。
10. 出行辅助指引：提供长途停靠提醒、周边宠物临时安置点位推荐服务。"""

examples_11 = """示例1：
输入：{"query": "我去便利店买个东西，狗留车上", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：夏天，车外温度35℃
输出：{"speak": "yes", "content": "[关心]好，空调给它保持24度，车窗锁上，屏幕我也锁了防它乱按，你放心去", "actions": [{"id": "ac_pet", "params": {"function": "air_conditioning", "mode": "cooling", "temperature": 24, "continuous": true}, "tool": "vehicle_basic_control"}, {"id": "lock_car", "params": {"function": "door_lock", "action": "lock"}, "tool": "vehicle_basic_control"}, {"id": "lock_screen", "params": {"function": "touch_screen", "action": "lock"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "我可能要半小时", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：车内有金毛，宠物模式已启动
输出：{"speak": "yes", "content": "[关心]半小时没问题，空调一直开着，我通过摄像头看着它，有啥异常马上通知你", "actions": [{"id": "oms_monitor", "params": {"target": "cabin", "mode": "pet_watch"}, "tool": "search_visual_info"}, {"id": "set_pet_goal", "params": {"action": "add", "goal": "持续监控宠物状态，异常立即推送车主APP"}, "tool": "goal_list_update"}]}

示例3：
输入：{"advisor": "OMS检测到宠物持续吠叫超过5分钟"}
env_info：车主已离车，宠物独处中
输出：{"speak": "no", "content": "", "actions": [{"id": "notify_owner", "params": {"function": "app_push", "message": "您的宠物持续叫了一段时间，可能有些焦虑，建议尽快回来看看"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("车内关怀·宠物关怀", logic_11 + "\n\n" + examples_11))

# ========== ROW 12: 车内环境·快速通风 ==========
logic_12 = """1. 获取信息：
用户是否有抽烟行为（通过车内PM2.5传感器检测）或明确表达通风需求（如"打开窗户通风""车内太闷了"）。
2. 判断通风触发条件：
当检测到抽烟行为或用户明确提出通风需求。
3. 控制车窗开启方式和程度：
根据实际情况（如车速、外界环境）合理调整车窗开启方式（如部分开启、全开）和程度。
4. 协调空调及天窗配合：
结合空调系统（如开启外循环）和天窗（如开启天窗进行换气），通过多系统协同工作以达到最佳通风效果。
5. 环境车况采集：自动感知车辆暴晒静置、车内起雾潮湿、多人乘车、驶出地库、途经隧道拥堵、长时间内循环等状态及空气质量数据。
6. 通风类型判定：依据场景匹配散热散味、除湿除雾、空气净化、尾气隔离对应的通风策略。
7. 行车工况校验：结合当前车速、外界气温与空气质量，限定通风强度和运行方式。
8. 多系统联动执行：自动切换内外循环、调节空调风量、适配风道方向，按需联动空气净化，不随意触发车窗天窗。
9. 动态自适应调节：行车提速自动弱化通风档位、降低风噪；环境变差自动锁定换气模式，阻隔尾气潮气侵入。
10. 状态闭环终止：监测车内空气、湿度达标后，自动退出强通风模式，恢复常规空调运行状态，并记忆本次使用偏好。"""

examples_12 = """示例1：
输入：{"query": "车里有点闷，通下风", "speaker": "主驾驶位", "zone": "主驾驶位"}
输出：{"speak": "yes", "content": "好", "actions": [{"id": "window_user", "params": {"function": "window", "position": "driver", "action": "open"}, "tool": "vehicle_basic_control"}, {"id": "window_diagonal", "params": {"function": "window", "position": "rear_right", "action": "open"}, "tool": "vehicle_basic_control"}, {"id": "ac_external", "params": {"function": "air_conditioning", "circulation": "external"}, "tool": "vehicle_basic_control"}, {"id": "sunroof_tilt", "params": {"function": "sunroof", "action": "tilt"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"advisor": "车内PM2.5传感器检测到异常升高"}
env_info：用户可能在抽烟
输出：{"speak": "no", "content": "", "actions": [{"id": "window_user", "params": {"function": "window", "position": "driver", "action": "open"}, "tool": "vehicle_basic_control"}, {"id": "window_diagonal", "params": {"function": "window", "position": "rear_right", "action": "open"}, "tool": "vehicle_basic_control"}, {"id": "ac_external", "params": {"function": "air_conditioning", "circulation": "external"}, "tool": "vehicle_basic_control"}]}

示例3：
输入：{"advisor": "通风5分钟后车内空气质量已恢复正常"}
env_info：通风模式运行中
输出：{"speak": "yes", "content": "[开心]空气好多了，窗户给你关上了", "actions": [{"id": "window_close", "params": {"function": "window", "position": "all", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "ac_internal", "params": {"function": "air_conditioning", "circulation": "internal"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("车内环境·快速通风", logic_12 + "\n\n" + examples_12))

# ========== ROW 13: 车内关怀·后排勿扰 ==========
logic_13 = """1. 获取需求：用户明确提出后排安静环境需求时，默认核心目标为后排保持相对安静，不打扰前排人员。

2. 营造安静环境：采取降低后排噪音源（如多媒体音量调至最低、关闭非必要提示音）、优化座椅隔音等措施，减少后排声音对前排的影响。

3. 儿童安全及状态查看方案：若后排有儿童，启用儿童安全锁确保乘车安全；通过车内摄像头实时监控儿童状态，或设置定时语音轻柔询问（避免打扰安静），并将状态信息同步至前排显示终端。

4. 平衡协调：在实施安静环境营造措施时，确保儿童安全监控功能正常运行，避免因过度追求安静而忽略儿童状态；监控过程中保持低干扰，不额外增加噪音，实现安静需求与安全保障的平衡。"""

examples_13 = """示例1：
输入：{"query": "后面孩子睡着了，别吵到他", "speaker": "主驾驶位", "zone": "主驾驶位"}
输出：{"speak": "yes", "content": "[轻柔]好，后排声音我都关了，灯也关了，你放心", "actions": [{"id": "rear_speaker_off", "params": {"function": "rear_speaker", "action": "mute"}, "tool": "vehicle_basic_control"}, {"id": "rear_light_off", "params": {"function": "rear_ambient_light", "action": "off"}, "tool": "vehicle_basic_control"}, {"id": "rear_reading_off", "params": {"function": "rear_reading_light", "action": "off"}, "tool": "vehicle_basic_control"}, {"id": "close_sunshade", "params": {"function": "rear_sunshade", "action": "close"}, "tool": "vehicle_basic_control"}, {"id": "rear_screen_off", "params": {"function": "rear_screen", "action": "off"}, "tool": "vehicle_basic_control"}, {"id": "child_lock", "params": {"function": "child_lock", "action": "on"}, "tool": "vehicle_basic_control"}, {"id": "rear_status_card", "params": {"function": "front_screen_card", "content": "rear_status", "action": "show"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "后面孩子醒了没", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：后排勿扰模式运行中
输出：{"speak": "yes", "content": "[轻柔]我看看", "actions": [{"id": "check_rear", "params": {"target": "rear_seat"}, "tool": "search_visual_info"}]}
（OMS返回结果后）
输出：{"speak": "yes", "content": "[轻柔]还在睡，安心开车", "actions": []}

示例3：
输入：{"advisor": "OMS检测到后排孩子已醒来并开始说话"}
env_info：后排勿扰模式运行中
输出：{"speak": "yes", "content": "[温柔]宝宝醒了，要不要恢复后排声音？", "actions": []}"""

rows_data.append(("车内关怀·后排勿扰", logic_13 + "\n\n" + examples_13))

# ========== ROW 14: 儿童娱乐·陪伴互动（系统触发） ==========
logic_14 = """1. 获取信息：
监测到儿童在后排的信号（如车内摄像头识别）。
确认车内外摄像头、语音及后排吸顶屏等多媒介手段是否可用。
2. 确定游戏形式：
结合多媒介手段，选择多样化的游戏形式，如车型品牌问答、颜色识别等。
3. 设计游戏内容：
注重互动性和趣味性，涵盖知识问答、音乐互动、视觉识别、寻宝等不同类型。
4. 分配娱乐空间：
利用后排独立的吸顶屏，为儿童提供专属的娱乐互动区域，所有游戏内容仅在后排屏幕显示，不向前排屏幕推送，避免干扰驾驶员驾驶；
所有游戏/互动内容需为法律或者车企授权的适龄儿童内容，无暴力、不良引导内容，符合儿童内容安全规范。
5. 互动执行与反馈：
游戏过程中采用轻柔的语音引导，避免过大音量打扰儿童和驾驶员；
每完成一个互动环节，给予儿童正向的语音鼓励，提升陪伴体验；
若检测到儿童情绪不佳/哭闹，立即暂停游戏内容，自动切换至安抚类音频/故事内容。"""

examples_14 = """示例1：
输入：{"advisor": "OMS检测到后排有儿童约5岁，行车中"}
env_info：后排吸顶屏可用
输出：{"speak": "yes", "content": "[开心]小朋友，我们玩个游戏吧！你看窗外，能找到几辆红色的车？", "actions": [{"id": "show_scoreboard", "params": {"function": "rear_screen", "content": "game_scoreboard"}, "tool": "vehicle_basic_control"}]}

示例2：
输入：{"query": "我要看动画", "speaker": "二排左", "zone": "二排左"}
env_info：儿童互动游戏进行中
输出：{"speak": "yes", "content": "[开心]好，给你放一集汪汪队", "actions": [{"id": "stop_game", "params": {"function": "rear_screen", "content": "game", "action": "stop"}, "tool": "vehicle_basic_control"}, {"id": "play_cartoon", "params": {"keyword": "汪汪队适龄动画", "target": "rear_screen"}, "tool": "search_and_control_short_video"}]}

示例3：
输入：{"query": "别太吵了", "speaker": "主驾驶位", "zone": "主驾驶位"}
env_info：后排儿童互动游戏进行中
输出：{"speak": "yes", "content": "好，我把后排声音小一点", "actions": [{"id": "rear_volume_down", "params": {"function": "rear_speaker", "action": "volume_down", "level": "low"}, "tool": "vehicle_basic_control"}]}"""

rows_data.append(("儿童娱乐·陪伴互动（系统触发）", logic_14 + "\n\n" + examples_14))

# Write all data
for i, (name, content) in enumerate(rows_data, 2):
    ws.cell(row=i, column=1, value=i-1).font = cell_font
    ws.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='top')
    ws.cell(row=i, column=2, value=name).font = cell_font
    ws.cell(row=i, column=2).alignment = Alignment(vertical='top', wrap_text=True)
    ws.cell(row=i, column=3, value=content).font = cell_font
    ws.cell(row=i, column=3).alignment = cell_align

# Freeze header row
ws.freeze_panes = 'A2'

output_path = "/Users/bytedance/Desktop/3.23/场景知识/执行逻辑_JSON示例.xlsx"
wb.save(output_path)
print(f"Done: {output_path}")
print(f"Total rows: {len(rows_data)}")
